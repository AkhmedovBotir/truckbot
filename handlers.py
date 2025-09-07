import csv
import io
from datetime import datetime
from telegram import Update, ReplyKeyboardRemove
from telegram.ext import ContextTypes, ConversationHandler
from telegram.constants import ParseMode

from database import db
from currency_api import currency_api
from keyboards import keyboards
from config import ADMIN_IDS

# Currency to country flag mapping
CURRENCY_FLAGS = {
    'USD': '🇺🇸', 'EUR': '🇪🇺', 'GBP': '🇬🇧', 'JPY': '🇯🇵', 'CHF': '🇨🇭',
    'CAD': '🇨🇦', 'AUD': '🇦🇺', 'NZD': '🇳🇿', 'SEK': '🇸🇪', 'NOK': '🇳🇴',
    'DKK': '🇩🇰', 'PLN': '🇵🇱', 'CZK': '🇨🇿', 'HUF': '🇭🇺', 'RUB': '🇷🇺',
    'CNY': '🇨🇳', 'KRW': '🇰🇷', 'SGD': '🇸🇬', 'HKD': '🇭🇰', 'INR': '🇮🇳',
    'BRL': '🇧🇷', 'MXN': '🇲🇽', 'ZAR': '🇿🇦', 'TRY': '🇹🇷', 'AED': '🇦🇪',
    'SAR': '🇸🇦', 'QAR': '🇶🇦', 'KWD': '🇰🇼', 'BHD': '🇧🇭', 'OMR': '🇴🇲',
    'JOD': '🇯🇴', 'LBP': '🇱🇧', 'EGP': '🇪🇬', 'MAD': '🇲🇦', 'TND': '🇹🇳',
    'DZD': '🇩🇿', 'LYD': '🇱🇾', 'SDG': '🇸🇩', 'ETB': '🇪🇹', 'KES': '🇰🇪',
    'UGX': '🇺🇬', 'TZS': '🇹🇿', 'ZMW': '🇿🇲', 'BWP': '🇧🇼', 'SZL': '🇸🇿',
    'LSL': '🇱🇸', 'NAD': '🇳🇦', 'MZN': '🇲🇿', 'AOA': '🇦🇴', 'XOF': '🇸🇳',
    'XAF': '🇨🇲', 'CDF': '🇨🇩', 'RWF': '🇷🇼', 'BIF': '🇧🇮', 'KMF': '🇰🇲',
    'DJF': '🇩🇯', 'SOS': '🇸🇴', 'ERN': '🇪🇷', 'ETB': '🇪🇹', 'STN': '🇸🇹',
    'CVE': '🇨🇻', 'GMD': '🇬🇲', 'GNF': '🇬🇳', 'LRD': '🇱🇷', 'SLL': '🇸🇱',
    'GHS': '🇬🇭', 'NGN': '🇳🇬', 'XPF': '🇵🇫', 'TOP': '🇹🇴', 'WST': '🇼🇸',
    'VUV': '🇻🇺', 'SBD': '🇸🇧', 'PGK': '🇵🇬', 'FJD': '🇫🇯', 'NPR': '🇳🇵',
    'PKR': '🇵🇰', 'LKR': '🇱🇰', 'BDT': '🇧🇩', 'MMK': '🇲🇲', 'THB': '🇹🇭',
    'LAK': '🇱🇦', 'KHR': '🇰🇭', 'VND': '🇻🇳', 'IDR': '🇮🇩', 'MYR': '🇲🇾',
    'PHP': '🇵🇭', 'BND': '🇧🇳', 'MOP': '🇲🇴', 'TWD': '🇹🇼', 'MNT': '🇲🇳',
    'KZT': '🇰🇿', 'UZS': '🇺🇿', 'KGS': '🇰🇬', 'TJS': '🇹🇯', 'AFN': '🇦🇫',
    'IRR': '🇮🇷', 'IQD': '🇮🇶', 'SYP': '🇸🇾', 'LBP': '🇱🇧', 'JOD': '🇯🇴',
    'ILS': '🇮🇱', 'PAL': '🇵🇸', 'EGP': '🇪🇬', 'LYD': '🇱🇾', 'TND': '🇹🇳',
    'DZD': '🇩🇿', 'MAD': '🇲🇦', 'SDG': '🇸🇩', 'ETB': '🇪🇹', 'SOS': '🇸🇴',
    'DJF': '🇩🇯', 'ERN': '🇪🇷', 'STN': '🇸🇹', 'CVE': '🇨🇻', 'GMD': '🇬🇲',
    'GNF': '🇬🇳', 'LRD': '🇱🇷', 'SLL': '🇸🇱', 'GHS': '🇬🇭', 'NGN': '🇳🇬',
    'XOF': '🇸🇳', 'XAF': '🇨🇲', 'CDF': '🇨🇩', 'RWF': '🇷🇼', 'BIF': '🇧🇮',
    'KMF': '🇰🇲', 'MGA': '🇲🇬', 'SCR': '🇸🇨', 'MUR': '🇲🇺', 'MVR': '🇲🇻',
    'NPR': '🇳🇵', 'PKR': '🇵🇰', 'LKR': '🇱🇰', 'BDT': '🇧🇩', 'MMK': '🇲🇲',
    'THB': '🇹🇭', 'LAK': '🇱🇦', 'KHR': '🇰🇭', 'VND': '🇻🇳', 'IDR': '🇮🇩',
    'MYR': '🇲🇾', 'PHP': '🇵🇭', 'BND': '🇧🇳', 'MOP': '🇲🇴', 'TWD': '🇹🇼',
    'MNT': '🇲🇳', 'KZT': '🇰🇿', 'UZS': '🇺🇿', 'KGS': '🇰🇬', 'TJS': '🇹🇯',
    'AFN': '🇦🇫', 'IRR': '🇮🇷', 'IQD': '🇮🇶', 'SYP': '🇸🇾', 'ILS': '🇮🇱',
    'PAL': '🇵🇸', 'XPF': '🇵🇫', 'TOP': '🇹🇴', 'WST': '🇼🇸', 'VUV': '🇻🇺',
    'SBD': '🇸🇧', 'PGK': '🇵🇬', 'FJD': '🇫🇯'
}

def format_currency_with_flag(currency_data):
    """Format currency data with country flag"""
    currency_code = currency_data.get('Ccy', '')
    currency_name = currency_data.get('CcyNm_EN', 'Unknown')
    rate = currency_data.get('Rate', '0')
    diff = currency_data.get('Diff', '0')
    
    # Get country flag
    flag = CURRENCY_FLAGS.get(currency_code, '🌍')
    
    # Format difference with arrow
    if diff and diff != '0':
        try:
            diff_float = float(diff)
            if diff_float > 0:
                diff_symbol = '📈 +'
            elif diff_float < 0:
                diff_symbol = '📉 -'
            else:
                diff_symbol = '➡️'
        except:
            diff_symbol = '➡️'
    else:
        diff_symbol = '➡️'
    
    return f"{flag} **{currency_name}** ({currency_code})\n💰 **Kurs:** {rate} so'm {diff_symbol} {diff}\n"

# Conversation states
WAITING_NAME, WAITING_PHONE, WAITING_TRACK_PHONE, WAITING_TRACK_CODE, WAITING_CHANNEL = range(5)

class BotHandlers:
    def __init__(self):
        pass

    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Start komandasi"""
        user_id = update.effective_user.id
        user = await db.get_user(user_id)
        
        if user:
            # Foydalanuvchi allaqachon ro'yxatdan o'tgan
            if user_id in ADMIN_IDS:
                await update.message.reply_text(
                    "👋 Xush kelibsiz, Admin!",
                    reply_markup=keyboards.admin_menu()
                )
            else:
                await update.message.reply_text(
                    f"👋 Xush kelibsiz, {user['fullname']}!",
                    reply_markup=keyboards.main_menu()
                )
        else:
            # Yangi foydalanuvchi - ro'yxatdan o'tishni boshlash
            await update.message.reply_text(
                "👋 Xush kelibsiz! Iltimos, to'liq ismingizni kiriting:",
                reply_markup=ReplyKeyboardRemove()
            )
            return WAITING_NAME

    async def handle_name(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Ro'yxatdan o'tish paytida ism kiritishni boshqarish"""
        context.user_data['fullname'] = update.message.text
        await update.message.reply_text(
            "📱 Iltimos, telefon raqamingizni ulashing:",
            reply_markup=keyboards.phone_request()
        )
        return WAITING_PHONE

    async def handle_phone(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Ro'yxatdan o'tish paytida telefon kiritishni boshqarish"""
        if update.message.contact:
            phone = update.message.contact.phone_number
        else:
            phone = update.message.text

        user_id = update.effective_user.id
        fullname = context.user_data.get('fullname', '')
        
        success = await db.add_user(user_id, fullname, phone)
        
        if success:
            if user_id in ADMIN_IDS:
                await update.message.reply_text(
                    "✅ Ro'yxatdan o'tish yakunlandi! Xush kelibsiz, Admin!",
                    reply_markup=keyboards.admin_menu()
                )
            else:
                await update.message.reply_text(
                    "✅ Ro'yxatdan o'tish yakunlandi! Botga xush kelibsiz!",
                    reply_markup=keyboards.main_menu()
                )
        else:
            await update.message.reply_text(
                "❌ Ro'yxatdan o'tish muvaffaqiyatsiz. Iltimos, qayta urinib ko'ring.",
                reply_markup=ReplyKeyboardRemove()
            )
        
        return ConversationHandler.END

    async def check_track_code(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Trek kodini tekshirish"""
        user_id = update.effective_user.id
        user = await db.get_user(user_id)
        
        if not user:
            await update.message.reply_text("❌ Iltimos, avval /start orqali ro'yxatdan o'ting")
            return

        phone = user['phone']
        user_data = await db.get_user_by_phone(phone)
        
        if user_data and user_data.get('track_code'):
            await update.message.reply_text(
                f"📦 Sizning trek kodingiz: `{user_data['track_code']}`",
                parse_mode=ParseMode.MARKDOWN
            )
        else:
            await update.message.reply_text("❌ Sizning raqamingiz uchun trek kod topilmadi.")

    async def check_currency(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Valyuta kursini tekshirish - admin tomonidan belgilangan valyutalarni ko'rsatish"""
        settings = await db.get_settings()
        selected_currencies = settings.get('selected_currencies', 'USD').split(',')
        selected_currencies = [curr.strip() for curr in selected_currencies if curr.strip()]
        
        if not selected_currencies:
            await update.message.reply_text("❌ Hozircha valyuta tanlanmagan. Admin bilan bog'laning.")
            return

        await update.message.reply_text(
            "💱 Valyutani tanlang:",
            reply_markup=keyboards.user_currencies_keyboard(selected_currencies)
        )

    async def my_profile(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Foydalanuvchi profilini ko'rsatish"""
        user_id = update.effective_user.id
        user = await db.get_user(user_id)
        
        if not user:
            await update.message.reply_text("❌ Iltimos, avval /start orqali ro'yxatdan o'ting")
            return

        profile_text = f"""
👤 **Mening profilim**

📝 To'liq ism: {user['fullname']}
📱 Telefon: {user['phone']}
📦 Trek kod: {user.get('track_code', 'Tayinlanmagan')}
📅 Ro'yxatdan o'tgan sana: {user['reg_date'][:10]}
        """
        
        await update.message.reply_text(profile_text, parse_mode=ParseMode.MARKDOWN)

    # Admin ishlovchilari
    async def currency_converter(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Valyuta konverteri menyusi"""
        print(f"Currency converter called by user: {update.effective_user.id}")
        print(f"Admin IDs: {ADMIN_IDS}")
        
        if update.effective_user.id not in ADMIN_IDS:
            await update.message.reply_text("❌ Ruxsat berilmagan.")
            return

        try:
            settings = await db.get_settings()
            print(f"Settings: {settings}")
            
            selected_currencies = settings.get('selected_currencies', 'USD').split(',')
            selected_currencies = [curr.strip() for curr in selected_currencies if curr.strip()]
            
            message = f"""💱 **Valyuta konverteri sozlamalari**

💰 **Tanlangan valyutalar:** {', '.join(selected_currencies)}
⏰ **Yuborish vaqti:** {settings.get('send_time', '09:00')}

📝 **Yuborish vaqtini o'zgartirish uchun:** HH:MM formatida vaqt yuboring (masalan: 14:30)

Quyidagi sozlamalarni tanlang:"""

            await update.message.reply_text(
                message,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=keyboards.currency_converter_menu()
            )
        except Exception as e:
            print(f"Error in currency_converter: {e}")
            await update.message.reply_text(f"❌ Xatolik: {str(e)}")

    async def users_list(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Foydalanuvchilar ro'yxatini ko'rsatish"""
        if update.effective_user.id not in ADMIN_IDS:
            await update.message.reply_text("❌ Ruxsat berilmagan.")
            return

        users = await db.get_all_users()
        
        if not users:
            await update.message.reply_text("👥 Foydalanuvchilar topilmadi.")
            return

        text = "👥 **Ro'yxatdan o'tgan foydalanuvchilar:**\n\n"
        for i, user in enumerate(users[:20], 1):  # Show first 20 users
            text += f"{i}. {user['fullname']} - {user['phone']}\n"
        
        if len(users) > 20:
            text += f"\n... va yana {len(users) - 20} ta foydalanuvchi"

        await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)

    async def track_code_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Trek kod tayinlash"""
        if update.effective_user.id not in ADMIN_IDS:
            await update.message.reply_text("❌ Ruxsat berilmagan.")
            return

        await update.message.reply_text(
            "📦 Trek kod tayinlash uchun telefon raqamini kiriting:",
            reply_markup=keyboards.cancel_keyboard()
        )
        return WAITING_TRACK_PHONE

    async def handle_track_phone(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Trek kod tayinlash uchun telefon raqamini boshqarish"""
        if update.message.text == "❌ Bekor qilish":
            await update.message.reply_text(
                "❌ Operatsiya bekor qilindi.",
                reply_markup=keyboards.admin_menu()
            )
            return ConversationHandler.END

        phone = update.message.text
        user = await db.get_user_by_phone(phone)
        
        if not user:
            await update.message.reply_text(
                "❌ Bu telefon raqami bilan foydalanuvchi topilmadi.",
                reply_markup=keyboards.admin_menu()
            )
            return ConversationHandler.END

        context.user_data['track_phone'] = phone
        await update.message.reply_text(
            f"📦 {user['fullname']} ({phone}) uchun trek kodini kiriting:",
            reply_markup=keyboards.cancel_keyboard()
        )
        return WAITING_TRACK_CODE

    async def handle_track_code(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Trek kod tayinlashni boshqarish"""
        if update.message.text == "❌ Bekor qilish":
            await update.message.reply_text(
                "❌ Operatsiya bekor qilindi.",
                reply_markup=keyboards.admin_menu()
            )
            return ConversationHandler.END

        track_code = update.message.text
        phone = context.user_data.get('track_phone')
        
        success = await db.update_track_code(phone, track_code)
        
        if success:
            await update.message.reply_text(
                f"✅ Trek kod `{track_code}` {phone} raqamiga tayinlandi",
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=keyboards.admin_menu()
            )
        else:
            await update.message.reply_text(
                "❌ Trek kod tayinlashda xatolik.",
                reply_markup=keyboards.admin_menu()
            )
        
        return ConversationHandler.END

    async def channels_management(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Kanallar boshqaruvi"""
        if update.effective_user.id not in ADMIN_IDS:
            await update.message.reply_text("❌ Ruxsat berilmagan.")
            return

        settings = await db.get_settings()
        channels = settings.get('channels', '').split(',')
        channels = [ch.strip() for ch in channels if ch.strip()]
        
        if channels:
            channels_text = "\n".join([f"• {channel}" for channel in channels])
            message = f"""📢 **Ulangan kanallar:**

{channels_text}

Yangi kanal qo'shish uchun kanal ID yoki username yuboring.
Kanal o'chirish uchun kanal ID yoki username yuboring."""
        else:
            message = """📢 **Ulangan kanallar yo'q**

Yangi kanal qo'shish uchun kanal ID yoki username yuboring.
Masalan: @mychannel yoki -1001234567890"""

        await update.message.reply_text(
            message,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=keyboards.cancel_keyboard()
        )
        return WAITING_CHANNEL

    async def handle_channel_input(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Kanal kiritishni boshqarish"""
        if update.message.text == "❌ Bekor qilish":
            await update.message.reply_text(
                "❌ Operatsiya bekor qilindi.",
                reply_markup=keyboards.admin_menu()
            )
            return ConversationHandler.END

        channel_input = update.message.text.strip()
        settings = await db.get_settings()
        channels = settings.get('channels', '').split(',')
        channels = [ch.strip() for ch in channels if ch.strip()]
        
        # Check if channel already exists
        if channel_input in channels:
            await update.message.reply_text(
                f"❌ Kanal {channel_input} allaqachon qo'shilgan.",
                reply_markup=keyboards.admin_menu()
            )
            return ConversationHandler.END
        
        # Add new channel
        channels.append(channel_input)
        success = await db.update_settings(channels=','.join(channels))
        
        if success:
            await update.message.reply_text(
                f"✅ Kanal {channel_input} muvaffaqiyatli qo'shildi!",
                reply_markup=keyboards.admin_menu()
            )
        else:
            await update.message.reply_text(
                "❌ Kanal qo'shishda xatolik.",
                reply_markup=keyboards.admin_menu()
            )
        
        return ConversationHandler.END

    async def handle_text_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle text messages for custom time input"""
        print(f"Text message handler called with text: {update.message.text}")
        
        if update.effective_user.id not in ADMIN_IDS:
            print("User is not admin, ignoring text message")
            return

        # Check if user is waiting for custom time input
        if context.user_data.get('waiting_for_custom_time'):
            context.user_data['waiting_for_custom_time'] = False
            
            time_input = update.message.text.strip()
            
            # Validate time format (HH:MM)
            try:
                from datetime import datetime
                datetime.strptime(time_input, '%H:%M')
            except ValueError:
                await update.message.reply_text(
                    "❌ Noto'g'ri vaqt formati. Iltimos, HH:MM formatida kiriting (masalan: 14:30)",
                    reply_markup=keyboards.cancel_keyboard()
                )
                context.user_data['waiting_for_custom_time'] = True
                return

            success = await db.update_settings(send_time=time_input)
            
            if success:
                # Update the scheduler with new time
                from scheduler import scheduler
                await scheduler.update_schedule()
                
                await update.message.reply_text(
                    f"✅ Yuborish vaqti {time_input} ga o'rnatildi",
                    reply_markup=keyboards.admin_menu()
                )
            else:
                await update.message.reply_text(
                    "❌ Vaqt saqlashda xatolik.",
                    reply_markup=keyboards.admin_menu()
                )
            return

        # Check if the message looks like a time input (HH:MM format)
        time_input = update.message.text.strip()
        try:
            from datetime import datetime
            datetime.strptime(time_input, '%H:%M')
            # If we get here, it's a valid time format
            success = await db.update_settings(send_time=time_input)
            
            if success:
                # Update the scheduler with new time
                from scheduler import scheduler
                await scheduler.update_schedule()
                
                await update.message.reply_text(
                    f"✅ Yuborish vaqti {time_input} ga o'rnatildi",
                    reply_markup=keyboards.admin_menu()
                )
            else:
                await update.message.reply_text(
                    "❌ Vaqt saqlashda xatolik.",
                    reply_markup=keyboards.admin_menu()
                )
        except ValueError:
            # Not a time format, ignore
            pass


    async def test_send_currencies(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Test currency sending"""
        if update.effective_user.id not in ADMIN_IDS:
            await update.message.reply_text("❌ Ruxsat berilmagan.")
            return

        # Import scheduler here to avoid circular import
        from scheduler import scheduler
        
        try:
            await scheduler.send_currency_update()
            await update.message.reply_text(
                "✅ Test xabari barcha foydalanuvchilar va kanallarga yuborildi!",
                reply_markup=keyboards.admin_menu()
            )
        except Exception as e:
            await update.message.reply_text(
                f"❌ Test yuborishda xatolik: {str(e)}",
                reply_markup=keyboards.admin_menu()
            )

    async def statistics(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Bot statistikasini ko'rsatish"""
        if update.effective_user.id not in ADMIN_IDS:
            await update.message.reply_text("❌ Ruxsat berilmagan.")
            return

        stats = await db.get_user_stats()
        settings = await db.get_settings()
        channels = settings.get('channels', '').split(',') if settings.get('channels') else []
        selected_currencies = settings.get('selected_currencies', 'USD').split(',')
        selected_currencies = [curr.strip() for curr in selected_currencies if curr.strip()]
        
        stats_text = f"""
📊 **Bot statistikasi**

👥 Jami foydalanuvchilar: {stats['total_users']}
🟢 Faol foydalanuvchilar (30 kun): {stats['active_users']}
📢 Ulangan kanallar: {len([ch for ch in channels if ch.strip()])}
💱 Tanlangan valyutalar: {', '.join(selected_currencies)}
⏰ Yuborish vaqti: {settings.get('send_time', '09:00')}
        """
        
        await update.message.reply_text(stats_text, parse_mode=ParseMode.MARKDOWN)

    async def export_users(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Foydalanuvchilarni Excel formatida eksport qilish"""
        if update.effective_user.id not in ADMIN_IDS:
            await update.message.reply_text("❌ Ruxsat berilmagan.")
            return

        users = await db.get_all_users()
        
        if not users:
            await update.message.reply_text("👥 Eksport qilish uchun foydalanuvchilar yo'q.")
            return

        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Excel workbook yaratish
            wb = Workbook()
            ws = wb.active
            ws.title = "Foydalanuvchilar"
            
            # Sarlavhalar
            headers = ['To\'liq ism', 'Telefon', 'Trek kod']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Ma'lumotlarni qo'shish
            for row, user in enumerate(users, 2):
                # Telefon raqamini to'g'ri formatda yozish
                phone = user['phone']
                if phone.startswith('+998'):
                    phone = phone[1:]  # + ni olib tashlash
                elif phone.startswith('998'):
                    pass  # allaqachon to'g'ri
                elif phone.startswith('8'):
                    phone = '998' + phone[1:]  # 8 ni 998 ga almashtirish
                elif phone.startswith('9'):
                    phone = '998' + phone  # 9 dan boshlansa 998 qo'shish
                
                ws.cell(row=row, column=1, value=user['fullname'])
                ws.cell(row=row, column=2, value=phone)
                ws.cell(row=row, column=3, value=user.get('track_code', ''))
            
            # Ustunlarni kengaytirish
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Excel faylini saqlash
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            filename = f"foydalanuvchilar_eksport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            await update.message.reply_document(
                document=excel_buffer,
                filename=filename,
                caption=f"📤 Foydalanuvchilar eksporti - {len(users)} ta foydalanuvchi"
            )
            
        except ImportError:
            # openpyxl yo'q bo'lsa, CSV formatida yuborish
            await update.message.reply_text("❌ Excel formatida eksport qilish uchun openpyxl kutubxonasi kerak. Avval CSV formatida yuborilmoqda...")
            
            # CSV yaratish
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['To\'liq ism', 'Telefon', 'Trek kod'])
            
            for user in users:
                # Telefon raqamini to'g'ri formatda yozish
                phone = user['phone']
                if phone.startswith('+998'):
                    phone = phone[1:]
                elif phone.startswith('998'):
                    pass
                elif phone.startswith('8'):
                    phone = '998' + phone[1:]
                elif phone.startswith('9'):
                    phone = '998' + phone
                
                writer.writerow([
                    user['fullname'],
                    phone,
                    user.get('track_code', '')
                ])
            
            csv_data = output.getvalue().encode('utf-8')
            filename = f"foydalanuvchilar_eksport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            await update.message.reply_document(
                document=io.BytesIO(csv_data),
                filename=filename,
                caption=f"📤 Foydalanuvchilar eksporti - {len(users)} ta foydalanuvchi"
            )

    # Callback ishlovchilari
    async def handle_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Callback so'rovlarini boshqarish"""
        query = update.callback_query
        await query.answer()
        
        data = query.data
        
        if data == "currency_converter":
            settings = await db.get_settings()
            selected_currencies = settings.get('selected_currencies', 'USD').split(',')
            selected_currencies = [curr.strip() for curr in selected_currencies if curr.strip()]
            
            message = f"""💱 **Valyuta konverteri sozlamalari**

💰 **Tanlangan valyutalar:** {', '.join(selected_currencies)}
⏰ **Yuborish vaqti:** {settings.get('send_time', '09:00')}

📝 **Yuborish vaqtini o'zgartirish uchun:** HH:MM formatida vaqt yuboring (masalan: 14:30)

Quyidagi sozlamalarni tanlang:"""
            
            await query.edit_message_text(
                message,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=keyboards.currency_converter_menu()
            )
        
        elif data == "choose_currency":
            currencies = await currency_api.get_available_currencies()
            settings = await db.get_settings()
            selected_currencies = settings.get('selected_currencies', 'USD').split(',')
            selected_currencies = [curr.strip() for curr in selected_currencies if curr.strip()]
            
            await query.edit_message_text(
                "💰 Valyutalarni tanlang (bir nechta tanlash mumkin):",
                reply_markup=keyboards.currency_selection_keyboard(currencies, selected_currencies)
            )
        
        elif data.startswith("toggle_currency_"):
            currency_code = data.replace("toggle_currency_", "")
            settings = await db.get_settings()
            selected_currencies = settings.get('selected_currencies', 'USD').split(',')
            selected_currencies = [curr.strip() for curr in selected_currencies if curr.strip()]
            
            if currency_code in selected_currencies:
                selected_currencies.remove(currency_code)
            else:
                selected_currencies.append(currency_code)
            
            # Update context for current selection
            context.user_data['temp_selected_currencies'] = selected_currencies
            
            currencies = await currency_api.get_available_currencies()
            await query.edit_message_text(
                "💰 Valyutalarni tanlang (bir nechta tanlash mumkin):",
                reply_markup=keyboards.currency_selection_keyboard(currencies, selected_currencies)
            )
        
        elif data == "save_currencies":
            selected_currencies = context.user_data.get('temp_selected_currencies', ['USD'])
            if not selected_currencies:
                selected_currencies = ['USD']
            
            await db.update_settings(selected_currencies=','.join(selected_currencies))
            await query.edit_message_text(
                f"✅ Valyutalar saqlandi: {', '.join(selected_currencies)}",
                reply_markup=keyboards.back_to_admin()
            )
        
        
        elif data == "test_send":
            # Execute test send immediately
            try:
                from scheduler import scheduler
                await scheduler.send_currency_update()
                await query.edit_message_text(
                    "✅ Test xabari barcha foydalanuvchilar va kanallarga yuborildi!",
                    reply_markup=keyboards.back_to_admin()
                )
            except Exception as e:
                await query.edit_message_text(
                    f"❌ Test yuborishda xatolik: {str(e)}",
                    reply_markup=keyboards.back_to_admin()
                )
        
        elif data == "check_schedule":
            # Check current schedule status
            try:
                from scheduler import scheduler
                jobs = scheduler.scheduler.get_jobs()
                job_info = []
                for job in jobs:
                    if job.id == 'currency_update':
                        job_info.append(f"⏰ Currency Update: {job.next_run_time}")
                
                if job_info:
                    message = "📅 **Joriy rejasi:**\n" + "\n".join(job_info)
                else:
                    message = "❌ Hech qanday reja topilmadi"
                
                await query.edit_message_text(
                    message,
                    reply_markup=keyboards.back_to_admin()
                )
            except Exception as e:
                await query.edit_message_text(
                    f"❌ Rejani tekshirishda xatolik: {str(e)}",
                    reply_markup=keyboards.back_to_admin()
                )
        
        
        
        elif data == "back_admin":
            settings = await db.get_settings()
            selected_currencies = settings.get('selected_currencies', 'USD').split(',')
            selected_currencies = [curr.strip() for curr in selected_currencies if curr.strip()]
            
            message = f"""💱 **Valyuta konverteri sozlamalari**

💰 **Tanlangan valyutalar:** {', '.join(selected_currencies)}
⏰ **Yuborish vaqti:** {settings.get('send_time', '09:00')}

📝 **Yuborish vaqtini o'zgartirish uchun:** HH:MM formatida vaqt yuboring (masalan: 14:30)

Quyidagi sozlamalarni tanlang:"""
            
            await query.edit_message_text(
                message,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=keyboards.currency_converter_menu()
            )
        
        elif data == "back_main":
            await query.edit_message_text(
                "👋 Asosiy menyuga qaytildi",
                reply_markup=keyboards.main_menu()
            )
        
        elif data.startswith("show_currency_"):
            currency_code = data.replace("show_currency_", "")
            currency_data = await currency_api.get_currency_by_code(currency_code)
            
            if currency_data:
                # Format currency with flag and trend
                currency_info = format_currency_with_flag(currency_data)
                message = f"""{currency_info}

📅 **Sana:** {currency_data.get('Date', 'Noma\'lum')}

📊 O'zbekiston Respublikasi Markaziy Banki"""
                await query.edit_message_text(message, parse_mode=ParseMode.MARKDOWN)
            else:
                await query.edit_message_text("❌ Valyuta ma'lumotlari topilmadi.")
        
        elif data == "show_all_currencies":
            # Show all selected currencies in one message
            settings = await db.get_settings()
            selected_currencies = settings.get('selected_currencies', 'USD').split(',')
            selected_currencies = [curr.strip() for curr in selected_currencies if curr.strip()]
            
            if not selected_currencies:
                await query.edit_message_text("❌ Hozircha valyuta tanlanmagan.")
                return
            
            # Collect all currency data
            currency_messages = []
            for currency_code in selected_currencies:
                currency_data = await currency_api.get_currency_by_code(currency_code)
                
                if currency_data:
                    currency_info = format_currency_with_flag(currency_data)
                    currency_messages.append(currency_info)
            
            if not currency_messages:
                await query.edit_message_text("❌ Valyuta ma'lumotlari topilmadi.")
                return
            
            # Combine all currencies into one message
            message = f"""📊 **Valyuta kurslari**
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{''.join(currency_messages)}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📅 **Sana:** {datetime.now().strftime('%d.%m.%Y')}
📊 **O'zbekiston Respublikasi Markaziy Banki**"""
            
            await query.edit_message_text(message, parse_mode=ParseMode.MARKDOWN)

# Global handlers instance
handlers = BotHandlers()